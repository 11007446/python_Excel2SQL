# -*- coding: UTF-8 -*-
import datetime
import openpyxl
import os
import shutil
import sql_insert_template
import configutil

# 1



def loadExcel():
    '''
    读取配置文件中待解析excel文件\文件列表,以及输出sql文件路径.
    解析Excel内容,生成Sql Insert语句执行,并保存到sql文件留档
    '''
    config = configutil.ConfigUtil()
    filepaths, outputpath_base, thesqltemplate, sheetname, startrow, sheetsize = config.getConfigString(
        'EXCELINPUTPATH'), config.getConfigString(
            'SQLOUTPUTPATH'), config.getConfigString(
                'SQLTEMPLATE'), config.getConfigString('SHEATNAME'), int(
                    config.getConfigString('PARSESTARTROW')), int(
                        config.getConfigString('SHEETSIZE'))
    print('解析EXCEL文件源: %s' % (filepaths))

    sqltemplate = getSqlTemplateByName(thesqltemplate)

    if type(filepaths) is list:
        # 文件路径中有多个Excel,逐一解析
        for filepath in filepaths:
            (filename, ext) = os.path.splitext(os.path.basename(filepath))
            outputpath_folder = '%s_%s' % (
                filename, str(datetime.datetime.now().strftime('%Y%m%d')))
            outputpath = outputpath_base + outputpath_folder + '/'
            e2s = Excel2sql(
                sqltemplate,
                [filepaths, sheetname, outputpath, startrow, sheetsize])
            e2s.parseWorkBook()
            pass
        pass
    elif type(filepaths) is str:
        # 若只有一个Excel文件，直接解析
        (filename, ext) = os.path.splitext(os.path.basename(filepaths))
        outputpath_folder = '%s_%s生成' % (
            filename, str(datetime.datetime.now().strftime('%Y%m%d')))
        outputpath = outputpath_base + outputpath_folder + '/'
        e2s = Excel2sql(
            sqltemplate,
            [filepaths, sheetname, outputpath, startrow, sheetsize])
        e2s.parseWorkBook()
        pass
    print('解析完毕', end='\n\n')


def getSqlTemplateByName(sql_template_name):
    template_obj = getattr(sql_insert_template, sql_template_name)()
    return template_obj


class Excel2sql(object):
    def __init__(self, sqltemplate, config):
        self.sqltemplate = sqltemplate
        self.config = config
        pass

    def parseWorkBook(self):
        # 文件路径
        # 需要解析Sheet名
        # 输出SQL文件路径
        # 起始解析行数
        # 单一SQL文件解析记录数
        filepath, parsesheetname, outputpath, startrow, sheetsize = self.config

        print('     解析EXCEL文件: %s' % (filepath))
        if os.path.exists(outputpath):
            shutil.rmtree(outputpath)
        os.makedirs(outputpath)

        wb = openpyxl.load_workbook(filepath)

        for sheet in wb.sheetnames:
            if (sheet == parsesheetname):
                # 只解析配置中指定sheet名的sheet，其他跳过
                print('         解析表%s开始' % (sheet))
                self.parseSheetData2SQL(wb[sheet], outputpath, startrow,
                                        sheetsize)
                print('         解析表%s完成' % (sheet), end='\n\n')
            pass
        print('     解析完毕', end='\n\n')

    def parseSheetData2SQL(self,
                           sheet,
                           output_path,
                           startRow=1,
                           sheetsize=5000):

        sheet_title = sheet.title
        columnpart, valuepart = self.sqltemplate.getSQL()
        maxRow, maxCol = sheet.max_row, sheet.max_column
        output_filename = sheet_title + '_' + str(
            datetime.datetime.now().strftime('%Y%m%d%H%M'))

        fw = open(
            output_path + output_filename + '.sql', 'w', encoding='utf-8')
        for loopindex, row in enumerate(
                sheet.iter_rows(
                    row_offset=startRow, max_row=maxRow - startRow)):
            rowlist = []
            for cell in row:
                cellvalue = cell.value
                if cellvalue is None:
                    # 若单元格为空,则添加空字符串
                    rowlist.append('')
                else:
                    rowlist.append(cellvalue)

            mainsql = columnpart + valuepart.format(rowlist)

            if (loopindex > 0 and loopindex % sheetsize == 0):
                fw.close()
                fw = open(
                    output_path + output_filename + '_' + str(loopindex) +
                    '.sql',
                    'w',
                    encoding='utf-8')

            fw.write(mainsql + '\n')
        fw.write('--%s数据共计%d条\n\n' % (sheet_title, loopindex))
        fw.close()
        # pass

    pass


if __name__ == "__main__":
    loadExcel()
    # templatetest = getSqlTemplateByName("SQL_insert_test")
    # column, value = templatetest.getSQL()
    # print("template column is %s" % column)
    # print("template value is %s" % value)
    pass
